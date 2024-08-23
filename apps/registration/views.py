import os
from rest_framework import viewsets, status
from rest_framework.response import Response
from openpyxl import Workbook, load_workbook
from django.conf import settings
from .models import Registration, Standart, Vip, SectorStandart, ReachStandart, SectorVip, ReachVip
from .serializers import (
    RegistrationSerializer,
    StandartSerializer,
    VipSerializer,
    SectorStandartSerializer,
    ReachStandartSerializer,
    SectorVipSerializer,
    ReachVipSerializer
)

class RegistrationViewSet(viewsets.ModelViewSet):
    queryset = Registration.objects.all()
    serializer_class = RegistrationSerializer

class StandartViewSet(viewsets.ModelViewSet):
    queryset = Standart.objects.all()
    serializer_class = StandartSerializer

class VipViewSet(viewsets.ModelViewSet):
    queryset = Vip.objects.all()
    serializer_class = VipSerializer

class SectorStandartViewSet(viewsets.ModelViewSet):
    queryset = SectorStandart.objects.all()
    serializer_class = SectorStandartSerializer

class ReachStandartViewSet(viewsets.ModelViewSet):
    queryset = ReachStandart.objects.all()
    serializer_class = ReachStandartSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        reach = serializer.save()  # Сохраняем данные в БД

        # Сохраняем данные в Excel файл
        self.save_to_excel(reach, 'Standart')

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def save_to_excel(self, reach, ticket_type):
        # Динамически строим путь к папке и файлу
        folder_path = os.path.join(settings.BASE_DIR, 'exel')
        file_path = os.path.join(folder_path, f'reach_{ticket_type}.xlsx')

        # Создаем папку, если она не существует
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        # Проверяем, существует ли файл; если нет, создаем новый
        if not os.path.exists(file_path):
            wb = Workbook()  # Создаем новый workbook
            sheet = wb.active
            # Заголовки
            headers = ['ФИО', 'Название компании', 'Сектор', 'Количество билетов', 'Email', 'Телефон']
            sheet.append(headers)
            wb.save(file_path)  # Сохраняем новый файл
        else:
            wb = load_workbook(file_path)  # Открываем существующий файл
            sheet = wb.active

        # Добавляем новые данные
        row = [
            reach.full_name,
            reach.name_company,
            reach.sector.name,
            reach.current,
            reach.email,
            reach.phone
        ]
        sheet.append(row)

        # Сохраняем файл
        wb.save(file_path)

class SectorVipViewSet(viewsets.ModelViewSet):
    queryset = SectorVip.objects.all()
    serializer_class = SectorVipSerializer

class ReachVipViewSet(viewsets.ModelViewSet):
    queryset = ReachVip.objects.all()
    serializer_class = ReachVipSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        reach = serializer.save()  # Сохраняем данные в БД

        # Сохраняем данные в Excel файл
        self.save_to_excel(reach, 'Vip')

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def save_to_excel(self, reach, ticket_type):
        # Динамически строим путь к папке и файлу
        folder_path = os.path.join(settings.BASE_DIR, 'exel')
        file_path = os.path.join(folder_path, f'reach_{ticket_type}.xlsx')

        # Создаем папку, если она не существует
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        # Проверяем, существует ли файл; если нет, создаем новый
        if not os.path.exists(file_path):
            wb = Workbook()  # Создаем новый workbook
            sheet = wb.active
            # Заголовки
            headers = ['ФИО', 'Название компании', 'Сектор', 'Количество билетов', 'Email', 'Телефон']
            sheet.append(headers)
            wb.save(file_path)  # Сохраняем новый файл
        else:
            wb = load_workbook(file_path)  # Открываем существующий файл
            sheet = wb.active

        # Добавляем новые данные
        row = [
            reach.full_name,
            reach.name_company,
            reach.sector.name,
            reach.current,
            reach.email,
            reach.phone
        ]
        sheet.append(row)

        # Сохраняем файл
        wb.save(file_path)