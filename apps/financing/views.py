import os
from rest_framework import viewsets, status
from rest_framework.views import APIView
from rest_framework.response import Response
from openpyxl import Workbook, load_workbook
from django.conf import settings
from .models import Financing, Image, Reach, Sector, LegalName, NameInfo, ImageForm
from apps.financing.serializers import (
    FinancingSerializer,
    ImageSerializer,
    ReachSerializer,
    SectorSerializer,
    LegalNameSerializer,
    ImageFormSerializer,
)

class FinancingViewSet(viewsets.ModelViewSet):
    queryset = Financing.objects.all()
    serializer_class = FinancingSerializer

class ImageViewSet(viewsets.ModelViewSet):
    queryset = Image.objects.all()
    serializer_class = ImageSerializer

class ImageFormViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ImageForm.objects.all()
    serializer_class = ImageFormSerializer

class ReachViewSet(viewsets.ModelViewSet):
    queryset = Reach.objects.all()
    serializer_class = ReachSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        reach = serializer.save()  # Сохраняем данные в БД

        # Сохраняем данные в Excel файл
        self.save_to_excel(reach)

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def save_to_excel(self, reach):
        # Динамически строим путь к папке и файлу
        folder_path = os.path.join(settings.BASE_DIR, 'exel')
        file_path = os.path.join(folder_path, 'bif.xlsx')

        # Создаем папку, если она не существует
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        try:
            # Попробуем открыть существующий файл
            wb = load_workbook(file_path)
            sheet = wb.active
        except FileNotFoundError:
            # Если файл не найден, создаем новый
            wb = Workbook()
            sheet = wb.active
            # Заголовки
            headers = ['ФИО', 'Название вашей компании', 'Юридическое название', 'Краткое описание', 'Сектор']
            sheet.append(headers)

        # Добавляем новые данные
        row = [
            reach.full_name,
            reach.name_company,
            reach.legal_name.name,
            reach.brief_description,
            reach.sector.name
        ]
        sheet.append(row)

        # Сохраняем файл
        wb.save(file_path)

class ExcelView(APIView):
    def get(self, request, format=None):
        # Динамически строим путь к файлу
        file_path = os.path.join(settings.BASE_DIR, 'exel', 'bif.xlsx')

        if not os.path.exists(file_path):
            return Response({"error": "Файл не найден."}, status=status.HTTP_404_NOT_FOUND)

        try:
            wb = load_workbook(filename=file_path)
        except Exception as e:
            return Response({"error": f"Не удалось открыть файл Excel: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        sheet = wb.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        
        return Response(data, status=status.HTTP_200_OK)

class SectorViewSet(viewsets.ModelViewSet):
    queryset = Sector.objects.all()
    serializer_class = SectorSerializer

class LegalNameViewSet(viewsets.ModelViewSet):
    queryset = LegalName.objects.all()
    serializer_class = LegalNameSerializer
