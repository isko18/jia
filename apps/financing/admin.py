from django.contrib import admin
from django.urls import path
from modeltranslation.admin import TranslationAdmin, TranslationTabularInline
from .models import Financing, Reach, Sector, LegalName, Image, NameInfo, ImageForm
from apps.financing.translation import *
from openpyxl import load_workbook
from django.http import HttpResponse
import os
from django.conf import settings
from django.shortcuts import redirect


class NameInfoInline(TranslationTabularInline):
    model = NameInfo
    extra = 1

class ImageFormAdmin(TranslationAdmin):
    inlines = [NameInfoInline]
    fieldsets = (
        ('Основные настройки', {
            'fields': ('image',),
        }),
    )

class FinancingAdmin(TranslationAdmin):
    fieldsets = (
        ('Основное', {
            'fields': ('image',),
        }),
        ('Русская версия', {
            'fields': ('title_ru',),
        }),
        ('Английская версия', {
            'fields': ('title_en',),
        }),
        ('Кыргызская версия', {
            'fields': ('title_ky',),
        }),
    )

class ReachAdmin(TranslationAdmin):
    fieldsets = (
        ('Основное', {
            'fields': ('full_name', 'name_company', 'legal_name', 'brief_description', 'sector'),
        }),
        ('Русская версия', {
            'fields': ('full_name_ru', 'name_company_ru', 'brief_description_ru'),
        }),
        ('Английская версия', {
            'fields': ('full_name_en', 'name_company_en', 'brief_description_en'),
        }),
        ('Кыргызская версия', {
            'fields': ('full_name_ky', 'name_company_ky', 'brief_description_ky'),
        }),
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('download-excel/', self.admin_site.admin_view(self.download_excel_file), name='download_excel'),
            path('view-excel/', self.admin_site.admin_view(self.view_excel_file), name='view_excel'),
            path('download-excel-standart/', self.admin_site.admin_view(self.download_excel_standart), name='download_excel_standart'),
            path('view-excel-standart/', self.admin_site.admin_view(self.view_excel_standart), name='view_excel_standart'),
            path('download-excel-vip/', self.admin_site.admin_view(self.download_excel_vip), name='download_excel_vip'),
            path('view-excel-vip/', self.admin_site.admin_view(self.view_excel_vip), name='view_excel_vip'),
        ]
        return custom_urls + urls

    def download_excel_file(self, request):
        file_path = os.path.join(settings.BASE_DIR, 'exel', 'bif.xlsx')

        if os.path.exists(file_path):
            wb = load_workbook(filename=file_path)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="bif.xlsx"'
            wb.save(response)
            return response
        else:
            self.message_user(request, "Файл не найден.", level="error")
            return redirect("..")

    def view_excel_file(self, request):
        file_path = os.path.join(settings.BASE_DIR, 'exel', 'bif.xlsx')

        if os.path.exists(file_path):
            wb = load_workbook(filename=file_path)
            sheet = wb.active

            # Формируем HTML для отображения содержимого файла
            html = '<html><body><h2>Содержимое Excel файла</h2><table border="1">'
            for row in sheet.iter_rows(values_only=True):
                html += '<tr>'
                for cell in row:
                    html += f'<td>{cell}</td>'
                html += '</tr>'
            html += '</table></body></html>'

            return HttpResponse(html)
        else:
            self.message_user(request, "Файл не найден.", level="error")
            return redirect("..")

    def download_excel_standart(self, request):
        file_path = os.path.join(settings.BASE_DIR, 'exel', 'reach_Standart.xlsx')

        if os.path.exists(file_path):
            wb = load_workbook(filename=file_path)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="reach_Standart.xlsx"'
            wb.save(response)
            return response
        else:
            self.message_user(request, "Файл не найден.", level="error")
            return redirect("..")

    def view_excel_standart(self, request):
        file_path = os.path.join(settings.BASE_DIR, 'exel', 'reach_Standart.xlsx')

        if os.path.exists(file_path):
            wb = load_workbook(filename=file_path)
            sheet = wb.active

            # Формируем HTML для отображения содержимого файла
            html = '<html><body><h2>Содержимое Excel файла (Standart)</h2><table border="1">'
            for row in sheet.iter_rows(values_only=True):
                html += '<tr>'
                for cell in row:
                    html += f'<td>{cell}</td>'
                html += '</tr>'
            html += '</table></body></html>'

            return HttpResponse(html)
        else:
            self.message_user(request, "Файл не найден.", level="error")
            return redirect("..")

    def download_excel_vip(self, request):
        file_path = os.path.join(settings.BASE_DIR, 'exel', 'reach_Vip.xlsx')

        if os.path.exists(file_path):
            wb = load_workbook(filename=file_path)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="reach_Vip.xlsx"'
            wb.save(response)
            return response
        else:
            self.message_user(request, "Файл не найден.", level="error")
            return redirect("..")

    def view_excel_vip(self, request):
        file_path = os.path.join(settings.BASE_DIR, 'exel', 'reach_Vip.xlsx')

        if os.path.exists(file_path):
            wb = load_workbook(filename=file_path)
            sheet = wb.active

            # Формируем HTML для отображения содержимого файла
            html = '<html><body><h2>Содержимое Excel файла (VIP)</h2><table border="1">'
            for row in sheet.iter_rows(values_only=True):
                html += '<tr>'
                for cell in row:
                    html += f'<td>{cell}</td>'
                html += '</tr>'
            html += '</table></body></html>'

            return HttpResponse(html)
        else:
            self.message_user(request, "Файл не найден.", level="error")
            return redirect("..")

class SectorAdmin(TranslationAdmin):
    fieldsets = (
        ('Русская версия', {
            'fields': ('name_ru',),
        }),
        ('Английская версия', {
            'fields': ('name_en',),
        }),
        ('Кыргызская версия', {
            'fields': ('name_ky',),
        }),
    )

class LegalNameAdmin(TranslationAdmin):
    fieldsets = (
        ('Русская версия', {
            'fields': ('name_ru',),
        }),
        ('Английская версия', {
            'fields': ('name_en',),
        }),
        ('Кыргызская версия', {
            'fields': ('name_ky',),
        }),
    )

admin.site.register(ImageForm, ImageFormAdmin)
admin.site.register(Financing, FinancingAdmin)
admin.site.register(Reach, ReachAdmin)
admin.site.register(Sector, SectorAdmin)
admin.site.register(LegalName, LegalNameAdmin)
admin.site.register(Image)
